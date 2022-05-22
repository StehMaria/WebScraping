import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


def planilha(Tabela):
    print('Iniciando verificação dos dados...')
    
    Tabela.get('CVSS_v20')
    cont = 0
    for i in Tabela.get('CVSS_v20'):
        #print(i)
        _,valores_v2,_ = i.split(' ')
        print(valores_v2)
        Tabela.get('CVSS_v20')[cont] = float(valores_v2)
        cont = cont+1

    print(Tabela.get('CVSS_v20'))
    cont=0
    for i in Tabela.get('CVSS_v31'):
        #print(i)
        _,valores_v3,_ = i.split(' ')
        print(valores_v3)
        Tabela.get('CVSS_v31')[cont] = float(valores_v3)
        cont = cont+1

    print(Tabela.get('CVSS_v31'))

    df = pd.DataFrame.from_dict(Tabela, orient ='columns') #dicionario para dataframe
    df = df.rename(columns={'CVE_ID':'CVE','CVSS_v31':'Severity_v3','CVSS_v20':
                        'Severity_v2','Descrição':'Current Description','Publicação':'NVD','Sistema':'Software/Sistema','Referencias':'References','Configuração':'Known'}) #renomear as colunas no dataframe
    
    df = df[['Software/Sistema',
                   'CVE',
                   'Current Description',
                  'Severity_v2','Severity_v3',
                  'References',
                  'Known',
                  'NVD',
                  'Link']] #organizar a posição das colunas
    print('Conversão para planilha finalizada!')
    #df.to_excel('Planilha.xlsx')
    df.to_excel('Vulnerabilidades.xlsx')
    print('Planilha criada com sucesso!')
    
    return df




def organizar(df):
    # Carregar dados para variável
    wb = load_workbook('Vulnerabilidades.xlsx')
    # Escolhe active sheet
    ws = wb.active
    # Deleta primeira coluna, que é somente índice
    ws.delete_cols(1)
    # Cabeçalho em negrito e fundo azul
    # Fill parameters
    my_fill = PatternFill(start_color='5399FF', 
                       end_color='5399FF',
                       fill_type='solid')
    # Bold Parameter
    my_font = Font(bold=True) ##tamanho e fonte
    # Formata o cabeçalho
    my_header = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1']
    for cell in my_header:
        ws[cell].fill = my_fill
        ws[cell].font = my_font
    # Formata as colunas e linhas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40
    for row in ws.iter_rows():  
        for cell in row:      
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    #alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    #cell = ws['C2']
    #ws.alignment = alignment
    # Salva o arquivo
    print('Planilha configurada')
    wb.save('Vulnerabilidades.xlsx')
    return 'Planilha salva'

#se há um valor maior que 8
def valores(Tabela):
    envio_2 = len([i for i in Tabela.get('CVSS_v20') if i > 8.0])
    print('Vulnerabilidades acima de 8 encontradas: ' + str(envio_2))
    envio_3 = len([i for i in Tabela.get('CVSS_v31') if i > 8.0])
    print('Vulnerabilidades acima de 8 encontradas: ' + str(envio_3))
    lista = []
    cont=0
    for i in Tabela.get('CVSS_v20'):
        if i > 8.0:
            lista.append(Tabela.get('CVE')[cont])
            lista.append(Tabela.get('CVSS_v20')[cont])
            lista.append(Tabela.get('CVSS_v31')[cont])
            lista.append(Tabela.get('Descrição')[cont])
            lista.append(Tabela.get('Publicação')[cont])    
            lista.append(Tabela.get('Link')[cont])
            lista.append(Tabela.get('Sistema')[cont])
            lista.append(Tabela.get('Referencias')[cont])
            lista.append(Tabela.get('Configuração')[cont])
        cont=cont+1
    cont=0
    for i in Tabela.get('CVSS_v31'):
        if i > 8.0:
            lista.append(Tabela.get('CVE')[cont])
            lista.append(Tabela.get('CVSS_v20')[cont])
            lista.append(Tabela.get('CVSS_v31')[cont])
            lista.append(Tabela.get('Descrição')[cont])
            lista.append(Tabela.get('Publicação')[cont])    
            lista.append(Tabela.get('Link')[cont])
            lista.append(Tabela.get('Sistema')[cont])
            lista.append(Tabela.get('Referencias')[cont])
            lista.append(Tabela.get('Configuração')[cont])
        cont=cont+1
    print(lista)
    return(lista,envio_3)